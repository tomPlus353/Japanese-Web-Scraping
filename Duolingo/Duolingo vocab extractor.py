#simple script to extract python data and write to an xlsm file
#uses openpyxl, os and of course duolingo

import openpyxl, duolingo, os
lingo = duolingo.Duolingo('tsull','myapipassword')
lingo.get_vocabulary()
vocab = lingo.get_vocabulary()
vocab = vocab["vocab_overview"]
workbook = openpyxl.load_workbook("Duolingo Data")
ws = workbook.active

#getting the titles
#remember structure List of Dictionarys with 14 key value pairs.
for row in ws.iter_rows(min_row=1, max_col=14, max_row=1):
	for cell in row:
		index = row.index(cell)
		columnTitle = list(vocab[0])[index]
		cell.value = columnTitle

#making a virtual table from a list of dictionaries
keyList = []
for row in ws.values:
	for value in row:
		keyList.append(value)

virTable = {}
for key in keyList:
	dataList = []
	for dictionary in vocab:
		dataList.append(dictionary[key])
	virTable[key] = dataList

## virtual table to populate the excel sheet
for category in virTable:
	dataList = virTable[category]
	for i in range(2, len(dataList)+1):
		listIndex = i - 2
		targetValue = dataList[listIndex]
		if type(targetValue) == list:
			targetValue = str(targetValue)
		ws.cell(i, keyList.index(category) + 1).value = targetValue

